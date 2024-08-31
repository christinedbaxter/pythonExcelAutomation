import shutil
import os
import pandas as pd
from openpyxl import load_workbook
import schedule
import time
import pytz
from datetime import datetime
from pytz import timezone
import logging

# Setup logging
logging.basicConfig(filename='script_log.log',
                    level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

def process_file(file_path, worksheet_name):
    try:
        # Load the file to a dataframe
        df = pd.read_csv(file_path, sep=',', header=0)
        
        # Load the main workbook and select the appropriate sheet
        workbook_name = 'mainFile.xlsx'
        wb = load_workbook(workbook_name)
        if worksheet_name in wb.sheetnames:
            ws = wb[worksheet_name]
        else:
            # If the worksheet doesn't exist, create it
            ws = wb.create_sheet(title=worksheet_name)
        
        # Write the new data to the main workbook
        newData = df.values.tolist()
        for row in newData:
            ws.append(row)
        
        # Save the workbook
        wb.save(filename=workbook_name)
        logging.info(f'Successfully processed and updated {file_path}')
    except Exception as e:
        logging.error(f'Error processing file {file_path}: {e}')

def job():
    try:
        # Define folder paths
        folders = ['folder1', 'folder2', 'folder3', 'folder4']
        processed_folder = 'processed/'

        # Get list of all processed files in the processed folder
        processed_files = [file for file in os.listdir(processed_folder) if file.endswith('.txt')]

        # Iterate over each folder
        for folder in folders:
            folder_path = os.path.join(os.getcwd(), folder)
            
            # Get list of all .txt files in the current folder
            input_files = [file for file in os.listdir(folder_path) if file.endswith('.txt')]
            
            for input_file in input_files:
                file_path = os.path.join(folder_path, input_file)
                
                # Extract DataNameTitle from the filename
                file_name = os.path.basename(file_path)
                data_name_title = file_name.split('_')[0]
                
                # Process the file if it hasn't been processed yet
                if file_name not in processed_files:
                    # Process the file and append data to the correct worksheet
                    process_file(file_path, data_name_title)
                    
                    # Move the processed file to the processed folder
                    shutil.move(file_path, os.path.join(processed_folder, file_name))
                    logging.info(f'Data from {file_name} added to main file')
                else:
                    logging.info(f'{file_name} already processed')
    except Exception as e:
        logging.error(f'Error during job execution: {e}')

# Schedule the job
def schedule_job():
    # Define the timezone
    tz = timezone('America/Chicago')
    
    # Define a wrapper function that handles timezone-aware scheduling
    def job_wrapper():
        now = datetime.now(tz=tz)
        if now.weekday() < 5:  # Check if today is a weekday (0=Monday, 4=Friday)
            job()

    # Schedule the job every day at 8:00 AM Central Time
    schedule.every().day.at("08:00").do(job_wrapper)

# Call the function to set up the schedule
schedule_job()

# Main loop to keep the script running
while True:
    schedule.run_pending()
    time.sleep(1)